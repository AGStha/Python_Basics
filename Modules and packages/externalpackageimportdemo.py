from colorama import init
from colorama import Fore
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime


# This is coloroma

init()
print(Fore.RED + "hello world")

# This is of package workbook

wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted

ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")
w=load_workbook('sample.xlsx')
print(w.sheetnames)
